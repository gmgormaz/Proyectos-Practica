import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font

df = pd.read_csv("Data/Measurements-Caida de tension.csv")
df2 = df
df = df.reset_index(drop=True)
df2 = df2.reset_index(drop=True)
prueba = ["Prueba de lazo sin disparos"]
prueba2 = ["Voltaje"]
df = df[df["Test Function"].isin(prueba)].copy()
df2 = df2[df2["Test Function"].isin(prueba2)].copy()

def pedir_in():
    while True:
        try:
            return int(input("In [A] (ej 16): ").strip())
        except:
            print("In inválido.")

n = len(df)

datos = {}
for i in range(1, n + 1):
    print(f"\n--- Circuito {i}/{n} ---")
    In = pedir_in()
    datos[i] = In


df["Circuito"] = ""
df["Z linea Ref. [Ω]"] = df["Primary Measurement"]
df["Primary Measurement"] = df2["Primary Measurement"].to_numpy()
df["Z linea [Ω]"] = ""
df["I nominal [A]"] = [datos[i] for i in range(1, len(df) + 1)]
df["ΔV [%]"] = ""

cols_base = [
    "Circuito",
    "Primary Measurement",
    "Z linea Ref. [Ω]",
    "Z linea [Ω]",
    "I nominal [A]",
    "ΔV [%]"
]   


conforme_sub = ["SI", "NO", "N/A", "Observación"]

out = df[cols_base].copy().reset_index(drop=True)

out = out.rename(columns={
    "Primary Measurement": "Voltaje"
})

cols_base = [
    "Circuito",
    "Voltaje",
    "Z linea Ref. [Ω]",
    "Z linea [Ω]",
    "I nominal [A]",
    "ΔV [%]"
]   



for s in conforme_sub:
    out[s] = ""

path = "Data\caida.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)

with pd.ExcelWriter(path, engine="openpyxl") as writer:
    out.to_excel(writer, index=False, header=False, startrow=2, sheet_name="Hoja1")

wb = openpyxl.load_workbook(path)
ws = wb["Hoja1"]

col_idx = 1

for c in cols_base:
    ws.cell(row=1, column=col_idx).value = c
    ws.cell(row=2, column=col_idx).value = ""
    col_idx += 1

start_conf = col_idx
for sub in conforme_sub:
    ws.cell(row=1, column=col_idx).value = "CONFORME"
    ws.cell(row=2, column=col_idx).value = sub
    col_idx += 1
end_conf = col_idx - 1

ws.merge_cells(start_row=1, start_column=start_conf, end_row=1, end_column=end_conf)

ws.freeze_panes = "A3"

bold = Font(bold=True)
max_col = ws.max_column
for c in range(1, max_col + 1):
    ws.cell(row=1, column=c).font = bold
    ws.cell(row=2, column=c).font = bold

wb.save(path)
