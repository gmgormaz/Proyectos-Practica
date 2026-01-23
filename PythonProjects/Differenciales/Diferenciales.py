import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from tkinter import Tk, filedialog
from pathlib import Path



def seleccionar_csv():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    archivo = filedialog.askopenfilename(
        title="Selecciona el CSV",
        filetypes=[("CSV", "*.csv"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return archivo

def nombre_salida_unico(path_xlsx: Path) -> Path:
    if not path_xlsx.exists():
        return path_xlsx
    i = 1
    while True:
        candidato = path_xlsx.with_stem(f"{path_xlsx.stem}_{i}")
        if not candidato.exists():
            return candidato
        i += 1

csv_path = seleccionar_csv()
if not csv_path:
    raise SystemExit("No se seleccionó ningún archivo.")

df = pd.read_csv(csv_path)



def limpiar_valor(x):
    if pd.isna(x):
        return x
    s = str(x)
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if not m:
        return x
    num = float(m.group())
    return int(num) if num.is_integer() else num

perm = [0, 1, 2, 3, 4, 9, 10, 7, 8, 5, 6, 11, 12, 13, 14, 15, 16]

df = df.iloc[::-1].reset_index(drop=True)

col = df.columns[9]
n = 8
df = df.iloc[: (len(df) // n) * n]

bottom8 = [" 0°", " 180°", "0°", "180°", "0° ", "180° ", " 0° ", " 180° "]
cat = {
    "ΔT a 5 x IΔn [ms]": [" 0°", " 180°"],
    "ΔT a 1 x IΔn [ms]": ["0°", "180°"],
    "ΔT a 1/2 x IΔn [ms]": ["0° ", "180° "],
    "ΔI a 1 x IΔn [ms]": [" 0° ", " 180° "],
}

simples_inicio = ["Dispositivo", "Referencia", "In [A]", "IΔn [mA]", "Clase"]
simple_despues = "Manual"
cat2 = "CONFORME"
sub_cat2 = ["SI", "NO", "N/A"]

vals8 = df[col].to_numpy().reshape(-1, n)
m = vals8.shape[0]

diferenciales = [f"Diferencial N°{i+1}" for i in range(m)]

simples_inicio_vals = np.full((m, len(simples_inicio)), "", dtype=object)
simple_despues_vals = np.full((m, 1), "", dtype=object)
cat2_vals = np.full((m, len(sub_cat2)), "", dtype=object)

values = np.concatenate([simples_inicio_vals, vals8, simple_despues_vals, cat2_vals], axis=1)

top = []
bottom = []

for h in simples_inicio:
    top.append(h)
    bottom.append("")

top8 = [next((k for k, cols in cat.items() if b in cols), "") for b in bottom8]
top += top8
bottom += bottom8

top.append(simple_despues)
bottom.append("")

for s in sub_cat2:
    top.append(cat2)
    bottom.append(s)

out = pd.DataFrame(values, columns=pd.MultiIndex.from_arrays([top, bottom]))
out = out.map(limpiar_valor)
out = out.iloc[:, perm]

out.loc[:, ("Dispositivo", "")] = diferenciales
out.loc[:, ("Clase", "")] = "AC"
out.loc[:, ("IΔn [mA]", "")] = "30"

cols_excluir = {
    ("ΔT a 1/2 x IΔn [ms]", "0° "),
    ("ΔT a 1/2 x IΔn [ms]", "180° "),
}

med_cols = [(top8[i], bottom8[i]) for i in range(8) if (top8[i], bottom8[i]) not in cols_excluir]

med_vals = out.loc[:, med_cols].apply(pd.to_numeric, errors="coerce")
ok = med_vals.apply(
    lambda r: (r.dropna() < 400).all(),
    axis=1
)


si = ("CONFORME", "SI")
no = ("CONFORME", "NO")
na = ("CONFORME", "N/A")

out.loc[:, si] = np.where(ok, "X", "")
out.loc[:, no] = np.where(ok, "", "X")
out.loc[:, na] = ""

csv_p = Path(csv_path)
base_dir = csv_p.parent.parent
path = base_dir / "Resultado_Diferenciales" / f"{csv_p.stem}_ordenado.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)

out.index.name = ""
out.to_excel(path, index=True)

wb = openpyxl.load_workbook(path)
ws = wb.active

ws["A1"].value = None
ws["A2"].value = None
ws.column_dimensions["A"].hidden = True

start_col = 2
end_col = start_col + out.shape[1] - 1

for c in range(start_col, end_col + 1):
    sub_cell = ws.cell(row=2, column=c)
    if isinstance(sub_cell, MergedCell):
        continue
    sub = sub_cell.value
    if sub is None or str(sub).strip() == "":
        sub_cell.value = None
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

row_extra = out.columns.nlevels + 1
ws.delete_rows(row_extra)
wb.save(path)

