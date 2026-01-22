import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font

df = pd.read_csv("Raw Data/Measurements-.csv")
prueba = ["Prueba de lazo sin disparos"]
df = df[df["Test Function"].isin(prueba)].copy()

ZS_MAX = {
    "B": {2:22, 4:11, 6:7.3, 10:4.4, 16:2.8, 20:2.2, 25:1.8, 32:1.4, 35:1.3, 40:1.1, 50:0.9, 63:0.7},
    "C": {2:11, 4:5.5, 6:3.65, 10:2.2, 16:1.4, 20:1.1, 25:0.9, 32:0.7, 35:0.65, 40:0.55, 50:0.45, 63:0.35},
    "D": {2:5.5, 4:2.8, 6:1.83, 10:1.1, 16:0.7, 20:0.55, 25:0.45, 32:0.34, 35:0.31, 40:0.27, 50:0.22, 63:0.17},
}

def pedir_in():
    while True:
        try:
            return int(input("In [A] (ej 16): ").strip())
        except:
            print("In inválido.")

def pedir_curva():
    while True:
        c = input("Curva (B/C/D): ").strip().upper()
        if c in ("B", "C", "D"):
            return c
        print("Curva inválida.")

n = len(df)

datos = {}
for i in range(1, n + 1):
    print(f"\n--- Circuito {i}/{n} ---")
    In = pedir_in()
    curva = pedir_curva()
    zsmax = ZS_MAX.get(curva, {}).get(In, np.nan)

    datos[i] = {"In": In, "curva": curva, "zsmax": zsmax}


df["Zs Max."] = ""
df["IPCC [A]"] = ""
df["In[A]/Curva/KA"] = ""
df["Circuito"] = ""

cols_base = [
    "Circuito",
    "In[A]/Curva/KA",
    "Primary Measurement",
    "Sub Measurement 2",
    "Zs Max.",
    "Sub Measurement 1",
    "Sub Measurement 3",
    "IPCC [A]",
]

df["In[A]/Curva/KA"] = [f'{datos[i]["curva"]}/{datos[i]["In"]}'for i in range(1, len(df) + 1)]

df["Zs Max."] = [datos[i]["zsmax"] for i in range(1, len(df) + 1)]


conforme_sub = ["SI", "NO", "N/A", "Observación"]

out = df[cols_base].copy().reset_index(drop=True)

pm = (
    out["Primary Measurement"]
    .astype(str)
    .str.replace(",", ".", regex=False)
    .str.extract(r"([-+]?\d*\.?\d+)")[0]
)

out["Primary Measurement"] = pd.to_numeric(pm, errors="coerce")
out["IPCC [A]"] = (220 * 1.06) / out["Primary Measurement"]


out = out.rename(columns={
    "Primary Measurement": "Zs [Ω]",
    "Sub Measurement 1": "PEFC [A]",
    "Sub Measurement 2": "Zl [Ω]",
    "Sub Measurement 3": "PSC [A]",
})

cols_base = [
    "Circuito",
    "In[A]/Curva/KA",
    "Zs [Ω]",
    "Zl [Ω]",
    "Zs Max.",
    "PEFC [A]",
    "PSC [A]",
    "IPCC [A]",
]



for s in conforme_sub:
    out[s] = ""

path = "Resultado_Prueba_Lazo\linea.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)

with pd.ExcelWriter(path, engine="openpyxl") as writer:
    out.to_excel(writer, index=False, header=False, startrow=2, sheet_name="Hoja1")

wb = openpyxl.load_workbook(path)
ws = wb["Hoja1"]

col_idx = 1

for c in cols_base:
    ws.cell(row=1, column=col_idx).value = c
    ws.cell(row=2, column=col_idx).value = ""
    col_idx += 1

start_conf = col_idx
for sub in conforme_sub:
    ws.cell(row=1, column=col_idx).value = "CONFORME"
    ws.cell(row=2, column=col_idx).value = sub
    col_idx += 1
end_conf = col_idx - 1

ws.merge_cells(start_row=1, start_column=start_conf, end_row=1, end_column=end_conf)

ws.freeze_panes = "A3"

bold = Font(bold=True)
max_col = ws.max_column
for c in range(1, max_col + 1):
    ws.cell(row=1, column=c).font = bold
    ws.cell(row=2, column=c).font = bold

wb.save(path)
