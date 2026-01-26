import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from tkinter import Tk, filedialog
from pathlib import Path


umbral = 1000

def seleccionar_csv():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    archivo = filedialog.askopenfilename(
        title="Selecciona el CSV",
        filetypes=[("CSV", "*.csv"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return archivo

def nombre_salida_unico(path_xlsx: Path) -> Path:
    if not path_xlsx.exists():
        return path_xlsx
    i = 1
    while True:
        candidato = path_xlsx.with_stem(f"{path_xlsx.stem}_{i}")
        if not candidato.exists():
            return candidato
        i += 1

csv_path = seleccionar_csv()
if not csv_path:
    raise SystemExit("No se seleccionó ningún archivo.")

df = pd.read_csv(csv_path)
df = df.iloc[::-1].reset_index(drop=True)

prueba2 = ['Prueba de aislamiento']
df2 = df[df['Test Function'].isin(prueba2)].copy()

if len(df2) == 0:
    raise SystemExit("No hay filas con 'Prueba de aislamiento' en el CSV.")

res = "Primary Measurement"
no = "Notas"
si = "Remark"

num2 = (
    df2[res]
    .astype(str)
    .str.extract(r'([-+]?\d*\.?\d+)')[0]
    .astype(float)
)

df2[no] = np.where(num2 < 0.5, "X", "")
df2[si] = np.where(num2 >= 0.5, "X", "")

df2 = df2.rename(columns={
    "Primary Measurement": "Resistencia",
    "Level A": "Circuito",
    "Remark": " SI ",
    "Notas": " NO "
})

Ce = "Level B"
alimentador = "Circuito"
num = pd.to_numeric(df2[alimentador], errors="coerce")
df2[alimentador] = np.where(num > umbral, "Al.", df2[Ce])

desired_columns = ["Circuito", "Configuración", "Resistencia", " SI ", " NO "]
df2 = df2[desired_columns].copy()



top = ["Circuito","Configuraci+on","Resistencia","CONFORME","CONFORME"]
bottom = [" "," "," ","SI","NO"]

out = pd.DataFrame(df2.values, columns=pd.MultiIndex.from_arrays([top, bottom]))

csv_p = Path(csv_path)
base_dir = csv_p.parent.parent
path = base_dir / "Resultado_Aislamiento" / f"{csv_p.stem}_aislamiento.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)
path = nombre_salida_unico(path)

out.index.name = ""
out.to_excel(path, index=True)

wb = openpyxl.load_workbook(path)
ws = wb.active

ws["A1"].value = None
ws["A2"].value = None
ws.column_dimensions["A"].hidden = True

start_col = 2
end_col = start_col + out.shape[1] - 1

for c in range(start_col, end_col + 1):
    sub_cell = ws.cell(row=2, column=c)
    if isinstance(sub_cell, MergedCell):
        continue
    sub = sub_cell.value
    if sub is None or str(sub).strip() == "":
        sub_cell.value = None
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

row_extra = out.columns.nlevels + 1
ws.delete_rows(row_extra)

wb.save(path)
