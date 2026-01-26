import os
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from pathlib import Path

def add_row(df, row_dict):
    df.loc[len(df)] = row_dict
    return df

def nombre_salida_unico(path_out: Path) -> Path:
    if not path_out.exists():
        return path_out
    i = 1
    while True:
        candidato = path_out.with_stem(f"{path_out.stem}_{i}")
        if not candidato.exists():
            return candidato
        i += 1

al = int(input("N° Alimentadores: "))
cir = int(input("N° Circuitos: "))

df = pd.DataFrame(columns=[
    "Circuito", "Configuración", "Resistencia", "SI", "NO", "Observación"
])

for i in range(0, cir):
    df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "N-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "L-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "L-N",  "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})

df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})

for i in range(0, al):
    df = add_row(df, {"Circuito": "", "Configuración": "N-PE",
                    "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    
    df = add_row(df, {"Circuito": "", "Configuración": "L-PE",
                    "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    
    df = add_row(df, {"Circuito": "", "Configuración": "L-N",  
                    "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
    
    df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "",
                       "SI": "", "NO": "", "Observación": ""})

top = ["Circuito", "Medición", "Medición", "CONFORME", "CONFORME", "CONFORME"]
bottom = ["N° Circuito", "Configuración", "Resistencia", "SI", "NO", "Observación"]
df.columns = pd.MultiIndex.from_arrays([top, bottom])

out_dir = Path("Tablas_vacias")
out_dir.mkdir(parents=True, exist_ok=True)
path = nombre_salida_unico(out_dir / "Tabla_aislamiento.xlsx")

df.index.name = ""
df.to_excel(path, index=True)
wb = openpyxl.load_workbook(path)
ws = wb.active

ws["A1"].value = None
ws["A2"].value = None
ws.column_dimensions["A"].hidden = True

start_col = 2
end_col = start_col + df.shape[1] - 1

for c in range(start_col, end_col + 1):
    sub_cell = ws.cell(row=2, column=c)
    if isinstance(sub_cell, MergedCell):
        continue
    sub = sub_cell.value
    if sub is None or str(sub).strip() == "":
        sub_cell.value = None
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

row_extra = df.columns.nlevels + 1
ws.delete_rows(row_extra)

wb.save(path)

print("Listo:", path)

