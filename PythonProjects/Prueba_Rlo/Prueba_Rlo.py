import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from tkinter import Tk, filedialog
from pathlib import Path


umbral = 1000

def seleccionar_csv():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    archivo = filedialog.askopenfilename(
        title="Selecciona el CSV",
        filetypes=[("CSV", "*.csv"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return archivo

def nombre_salida_unico(path_xlsx: Path) -> Path:
    if not path_xlsx.exists():
        return path_xlsx
    i = 1
    while True:
        candidato = path_xlsx.with_stem(f"{path_xlsx.stem}_{i}")
        if not candidato.exists():
            return candidato
        i += 1

csv_path = seleccionar_csv()
if not csv_path:
    raise SystemExit("No se seleccionó ningún archivo.")

df0 = pd.read_csv(csv_path)
df0 = df0.iloc[::-1].reset_index(drop=True)

prueba = ['Prueba de continuidad']
df = df0[df0['Test Function'].isin(prueba)].copy()

if len(df) == 0:
    raise SystemExit("No hay filas con 'Prueba de continuidad' en el CSV.")

res = "Primary Measurement"
no = "Notas"
si = "Remark"

num2 = (
    df[res]
    .astype(str)
    .str.extract(r'([-+]?\d*\.?\d+)')[0]
    .astype(float)
)

df[no] = np.where(num2 > 2, "X", "")
df[si] = np.where(num2 <= 2, "X", "")

rows = df.to_dict("records")
salida = []
i = 0

while i < len(rows):
    a = rows[i]

    if i + 1 < len(rows) and a["Configuración"] == rows[i+1]["Configuración"]:
        b = rows[i+1]
        out = {**{f"{k}_B": v for k, v in a.items()},
               **{f"{k}_A": v for k, v in b.items()}}
        i += 2
    else:
        out = {f"{k}_B": v for k, v in a.items()}
        out.update({f"{k}_A": np.nan for k in a.keys()})
        i += 1

    salida.append(out)

df_final = pd.DataFrame(salida)

contexto = ["Configuración", "Test Function", "Level A", "Level C"]
for c in contexto:
    df_final[f"{c}_A"] = df_final[f"{c}_A"].fillna(df_final[f"{c}_B"])

desired_columns = [
    "Configuración_A","Test Function_A","Level A_A","Level C_A",
    "Primary Measurement_A","Primary Measurement_B", "Remark_A", "Notas_A"
]
df = df_final[desired_columns].copy()
df = df.reset_index(drop=True)

col_circ = "Level C_A"
col_cfg  = "Configuración_A"

cargas = []
circuito_prev = None
cfg_prev = None
cfg_block = 0

for j in range(len(df)):
    circuito = str(df.at[j, col_circ]).strip()
    cfg = str(df.at[j, col_cfg]).strip()

    if j == 0 or circuito != circuito_prev:
        cfg_block = 0
        cfg_prev = None

    if cfg_prev is None or cfg != cfg_prev:
        cfg_block += 1

    carga = ((cfg_block - 1) // 3) + 1
    cargas.append(carga)

    circuito_prev = circuito
    cfg_prev = cfg

df["Carga"] = cargas

df = df.rename(columns={
    "Primary Measurement_A": "Resistencia (-)",
    "Primary Measurement_B": "Resistencia (+)",
    "Level C_A": "N° Circuito",
    "Level A_A": "Alimentador",
    "Test Function_A": "Prueba",
    "Configuración_A": "Configuración",
    "Remark_A": " SI ",
    "Notas_A": " NO "})

alimentador = "Alimentador"
cir = "N° Circuito"
carga_ = "Carga"

num = pd.to_numeric(df[alimentador], errors="coerce")
df[cir] = np.where(num > umbral, df[alimentador], df[cir])
df[carga_] = np.where(num > umbral, "Al.", df[carga_])

desired_columns = ["N° Circuito", "Carga", "Configuración",
                   "Resistencia (-)", "Resistencia (+)", " SI ", " NO "]
df_out = df[desired_columns].copy()


top = ["N° Circuito","Carga","Configuracion","Resistencia (-)","Resistencia (+)","CONFORME","CONFORME"]
bottom = ["","","","","","SI","NO"]

out = pd.DataFrame(df_out.values, columns=pd.MultiIndex.from_arrays([top, bottom]))

csv_p = Path(csv_path)
base_dir = csv_p.parent.parent
path = base_dir / "Resultado_Continuidad" / f"{csv_p.stem}_continuidad.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)
path = nombre_salida_unico(path)

out.index.name = ""
out.to_excel(path, index=True)

wb = openpyxl.load_workbook(path)
ws = wb.active

ws["A1"].value = None
ws["A2"].value = None
ws.column_dimensions["A"].hidden = True

start_col = 2
end_col = start_col + out.shape[1] - 1

for c in range(start_col, end_col + 1):
    sub_cell = ws.cell(row=2, column=c)
    if isinstance(sub_cell, MergedCell):
        continue
    sub = sub_cell.value
    if sub is None or str(sub).strip() == "":
        sub_cell.value = None
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

row_extra = out.columns.nlevels + 1
ws.delete_rows(row_extra)

wb.save(path)
